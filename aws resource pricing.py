import awspricing

ec2_offer = awspricing.offer('AmazonEC2')

ec2_offer.search_skus(
  instance_type='c4.large',
  location='US East (N. Virginia)',
  operating_system='Linux',
)  # {'4C7N4APU9GEUZ6H6', 'MBQPYDJSY3BY84BH', 'MDKVAJXMJGZFDJUE'}

ec2_offer.reserved_hourly(
  'c4.xlarge',
  operating_system='Linux',
  lease_contract_length='3yr',
  offering_class='convertible',
  purchase_option='Partial Upfront',
  region='us-east-1'
)  # 0.10845205479452055 , WSJ 1 

rds_offer = awspricing.offer('AmazonRDS')

rds_offer.search_skus(
  instance_type='db.m4.large',
  location='US East (N. Virginia)',
  database_engine='MySQL',
  license_model='No license required',
  deployment_option='Multi-AZ'
) # {'QPZNR6MYN432XTPU'}

rds_offer.ondemand_hourly(
  'db.m4.large',
  'MySQL',
  license_model='No license required',
  deployment_option='Multi-AZ',
  region='us-east-1'
) # 0.35

#!/usr/bin/env bash

# Run all AWS CLI commands and verify that 1 result is returned
# Make sure to run: terraform apply (with `local.debug_output = true` to populate aws_cli_commands)

set +e # do not exit on bad exit code

aws_cli_commands=$(terraform output -json aws_cli_commands | jq -r '.[] | gsub("[\\n\\t]"; "")')

while IFS=$'\n' read -ra commands; do
  for one in "${commands[@]}"; do
    printf '%*s\n' 58 | tr ' ' '-'
    echo $one
    result=$(/bin/bash -c "$one")
    exit_code=$?

    if [ $exit_code -ne 0 ]; then
      echo "Error!"
      continue
    fi

    products=$(echo $result | jq -r '.PriceList | length')

    if [ $products -eq 0 ]; then
      echo "0 products found. Expected 1!"
    elif [ $products -ne 1 ]; then
      echo "$products products found. Expected 1!"
      echo $result | jq -r '.PriceList[0] | fromjson | .product'
      echo $result | jq -r '.PriceList[1] | fromjson | .product'
    else
      echo "OK!"
      echo $result | jq -r '.PriceList[0] | fromjson | .terms.OnDemand[].priceDimensions[].pricePerUnit.USD'
      echo $result | jq -r '.PriceList[0] | fromjson | .product'
    fi

  done
done <<< "$aws_cli_commands"


#! /usr/local/bin/python3

import argparse
from openpyxl import load_workbook
import boto3
import json
import re

parser = argparse.ArgumentParser(description="Example: python aws-ec2-pricing.py -f <file>.xlsx -w <worksheet> -r us-west-2 -i m5 -v gp2")
parser.add_argument('-f', '--file', help='File name of the Excel workbook', required=True)
parser.add_argument('-w', '--ws', help='Worksheet or tab name of the Excel workbook', required=True)
parser.add_argument('-r', '--reg', help='AWS region, i.e. us-west-2', required=True)
parser.add_argument('-i', '--inst', help='EC2 supported values: all, t3, t2, m5, m4', required=True)
parser.add_argument('-v', '--vol', help='EBS supported values: gp2, st1', required=True)
args = parser.parse_args()
#print (args.file)
#print (args.ws)
#print (args.reg)
#print (args.inst)
#print (args.vol)

# Excel workbook and worksheet variables
wb = load_workbook(args.file) # The name of the workbook file to open passed from argparse
ws = wb[args.ws] # The name of the worksheet to manipulate
ws.cell(row=(1), column=6).value = 'Instance CPU'
ws.cell(row=(1), column=7).value = 'Instance RAM'
ws.cell(row=(1), column=8).value = 'Instance Final'
ws.cell(row=(1), column=9).value = 'Instance Unit Cost'
ws.cell(row=(1), column=10).value = 'Instance Daily Cost'
ws.cell(row=(1), column=11).value = 'EBS Unit Cost'
ws.cell(row=(1), column=12).value = 'EBS Daily Cost'

# Static AWS Pricing API variables
tenancy = "Shared" # Shared, Dedicated
preInstalledSw = "NA" # NA, SQL Ent, SQL Std, SQL Web
licenseModel = "No License required" # Bring your own license, NA, No License required

# AWS Regions excluding China and US Government
regionVar = [['ap-south-1','Asia Pacific (Mumbai)'],
	     	 ['ap-northeast-3','Asia Pacific (Osaka-Local)'],
       	     ['ap-northeast-2','Asia Pacific (Seoul)'],
             ['ap-southeast-1','Asia Pacific (Singapore)'],
	         ['ap-southeast-2','Asia Pacific (Sydney)'],
	         ['ap-northeast-1','Asia Pacific (Tokyo)'],
	         ['ca-central-1','Canada (Central)'],
	         ['eu-central-1','EU (Frankfurt)'],
	         ['eu-west-1','EU (Ireland)'],
	         ['eu-west-2','EU (London)'],
	         ['eu-west-3','EU (Paris)'],
	         ['eu-north-1','EU (Stockholm)'],
	         ['sa-east-1','South America (Sao Paulo)'],
	         ['us-east-1','US East (N. Virginia)'],
	         ['us-east-2','US East (Ohio)'],
	         ['us-west-1','US West (N. California)'],
	         ['us-west-2','US West (Oregon)']]

# Dynamic AWS Pricing API location variable
for x in regionVar:
	#print (args.reg)

	if (args.reg) == (x[0]):
		location = (x[1])
		#print (location)

# Dynamic AWS Pricing API volumeType variable
# EBS Volume Types: Cold HDD, General Purpose, Magnetic, Provisioned IOPS, Throughput Optimized HDD.
if (args.vol) == 'gp2':
	volumeType = 'General Purpose'
elif (args.vol) == 'st1':
	volumeType = 'Throughput Optimized HDD'
else:
	exit("error: no valid volume type: gp2, st1")
#print (volumeType)

if (args.inst) == 'all':
	instanceVar = "all"
elif (args.inst) == 't3':
	instanceVar = [['t3.nano',2,0.5],
				   ['t3.micro',2,1],
				   ['t3.small',2,2],
				   ['t3.medium',2,4],
				   ['t3.large',2,8],
				   ['t3.xlarge',4,16],
				   ['t3.2xlarge',8,32]]
elif (args.inst) == 't2':
	instanceVar = [['t2.nano',1,0.5],
				   ['t2.micro',1,1],
				   ['t2.small',1,2],
				   ['t2.medium',2,4],
				   ['t2.large',2,8],
				   ['t2.xlarge',4,16],
				   ['t2.2xlarge',8,32]]
elif (args.inst) == 'm5':
	instanceVar = [['m5.large',2,8],
				   ['m5.xlarge',4,16],
				   ['m5.2xlarge',8,32],
				   ['m5.4xlarge',16,64],
				   ['m5.12xlarge',48,192],
				   ['m5.24xlarge',96,384]]
elif (args.inst) == 'm4':
	instanceVar = [['m4.large',2,8],
				   ['m4.xlarge',4,16],
				   ['m4.2xlarge',8,32],
				   ['m4.4xlarge',16,64],
				   ['m4.10xlarge',40,160],
				   ['m4.16xlarge',64,256]]
else:
	exit("error: no valid instance family: all, t3, t2, m5, m4")

def rowRange():
	global row_count
	row_count = 0
	for x in range(2,10000):

		if ws.cell(row=x,column=1).value != None:
			#print(ws.cell(row=x,column=13).value)
			row_count = row_count + 1
		else:
			break

	row_count = row_count + 2
	#print (row_count)

def assignInstance():
	global row_count
	global instanceError
	instanceError = 0
	#print (instanceError)

	for x in range(2, row_count):
		sourceCpu = (ws.cell(row=(x), column=2).value)

		for y in instanceVar:
			instanceCpu = (y[1])

			if (sourceCpu) <= (instanceCpu):
				selectInstance = (y[0])
				#print (sourceCpu,instanceCpu)
				ws.cell(row=(x), column=6).value = (selectInstance)

				break

	for x in range(2, row_count):
		sourceRam = (ws.cell(row=x, column=3).value)/1024

		for y in instanceVar:
			instanceRam = (y[2])

			if (sourceRam-2) <= (instanceRam):
				selectInstance = (y[0])
				#print (sourceRam,instanceRam)
				ws.cell(row=(x), column=7).value = (selectInstance)

				break

	for x in range(2, row_count):
		i1 = (ws.cell(row=x, column=6).value)
		i2 = (ws.cell(row=x, column=7).value)
		#print (i1)
		#print (i2)

		for y in instanceVar:
			instanceType = (y[0])

			if (i1) == (instanceType):
				i1ram = (y[2])
				#print (i1ram)

		for y in instanceVar:
			instanceType = (y[0])

			if (i2) == (instanceType):
				i2ram = (y[2])
				#print (i2ram)

		if i2 != None:
			if (i1ram) <= (i2ram):
				selectInstance = i2
				ws.cell(row=(x), column=8).value = (selectInstance)
			else:
				selectInstance = i1
				ws.cell(row=(x), column=8).value = (selectInstance)
		else:
			instanceError = instanceError + 1
			#print (instanceError)

	if instanceError == 0:
		awsPricing()
	else:
		print("Error: Hosts source RAM is too large for {} instance(s)!".format(instanceError))
		print("Manually assign 'Instance Final' and re-run script using -i all")

def awsPricing():
	pricing = boto3.client('pricing')
	global row_count
	for x in range(2, row_count):
		# Dynamic AWS Pricing API instanceType variable
		instanceType = (ws.cell(row=x, column=8).value)

		# Dynamic AWS Pricing API operatingSystem variable
		# Operating System Types: Generic, Linux, NA, RHEL, SUSE, Windows
		operatingSystemString = (ws.cell(row=x, column=5).value)
		operatingSystemMatch = re.search( r'.*(Windows|Red Hat|SUSE).*', operatingSystemString, re.S)

		if operatingSystemMatch:
			if operatingSystemMatch.group(1) == 'Windows':
				operatingSystem = "Windows"
			elif operatingSystemMatch.group(1) == 'Red Hat':
				operatingSystem = "RHEL"
			elif operatingSystemMatch.group(1) == 'SUSE':
				operatingSystem = "SUSE"
			else:
				operatingSystem = "Linux"
		else:
			operatingSystem = "Linux"

		#print (operatingSystem)

		# AWS Pricing API EC2 Instance Attribute Match
		instanceData = pricing.get_products(
    		ServiceCode='AmazonEC2',
     		Filters = [
        		{'Type' :'TERM_MATCH', 'Field':'location',        'Value':location		  },
         		{'Type' :'TERM_MATCH', 'Field':'tenancy',         'Value':tenancy         },
         		{'Type' :'TERM_MATCH', 'Field':'instanceType',    'Value':instanceType    },
         		{'Type' :'TERM_MATCH', 'Field':'operatingSystem', 'Value':operatingSystem },
         		{'Type' :'TERM_MATCH', 'Field':'preInstalledSw',  'Value':preInstalledSw  },
         		{'Type' :'TERM_MATCH', 'Field':'licenseModel',    'Value':licenseModel    },

     		],
     		#MaxResults=100
		)

    	# Return output from pricing API
		#instanceString = str(instanceData)
		#print (instanceString)

		# When the Capacity Reservation is active, you are charged the equivalent On-Demand rate whether you run the instances or not.
		# If you do not use the reservation, this shows up as unused reservation on your EC2 bill.
		# When you run an instance that matches the attributes of a reservation, you just pay for the instance and nothing for the reservation.

		#AllocatedCapacityReservation
		#UnusedCapacityReservation
		#Used = OnDemand Pricing

		for instanceVal in instanceData["PriceList"]:
			instanceValJson=json.loads(instanceVal)
			if("OnDemand" in instanceValJson["terms"] and len(instanceValJson["terms"]["OnDemand"]) > 0):
				for onDemandValues in instanceValJson["terms"]["OnDemand"].keys():
					for priceDimensionValues in instanceValJson["terms"]["OnDemand"][onDemandValues]["priceDimensions"]:
						if("Used" in instanceValJson["product"]["attributes"]["capacitystatus"]):
							instancePrice = (instanceValJson["terms"]["OnDemand"][onDemandValues]["priceDimensions"][priceDimensionValues]["pricePerUnit"])

		instanceString = str(instancePrice)
    	#print (instanceString)

		instanceUnitPrice = re.search( r'(\d{1,10}\.\d{1,10})', instanceString, re.S)
    	#print (instanceUnitPrice)

		if instanceUnitPrice:
			instanceUnitPrice = instanceUnitPrice.group(1)
			ws.cell(row=(x), column=9).value = (instanceUnitPrice)
			#print ("Instance Unit Price:   " + (instanceUnitPrice))

			instanceDailyPrice = (float(instanceUnitPrice) * 24)
			#instanceDailyPrice = round(instanceDailyPrice,10)
			ws.cell(row=(x), column=10).value = (instanceDailyPrice)
			#print ("Instance Daily Price:  " + str(instanceDailyPrice))
		else:
			print ("No instance price match!")

		# EBS Storage Attribute Match
		storageData = pricing.get_products(
    		ServiceCode='AmazonEC2',
    		Filters = [
        		{'Type' :'TERM_MATCH', 'Field':'location',        'Value':location		  },
         		{'Type' :'TERM_MATCH', 'Field':'volumeType',      'Value':volumeType 	  },
     		],
     		#MaxResults=100
		)

		# Return output from pricing API
		storageString = str(storageData)
		#print (storageString)

		for storageVal in storageData["PriceList"]:
			storageValJson=json.loads(storageVal)
			if("OnDemand" in storageValJson["terms"] and len(storageValJson["terms"]["OnDemand"]) > 0):
				for onDemandValues in storageValJson["terms"]["OnDemand"].keys():
					for priceDimensionValues in storageValJson["terms"]["OnDemand"][onDemandValues]["priceDimensions"]:
						storagePrice = (storageValJson["terms"]["OnDemand"][onDemandValues]["priceDimensions"][priceDimensionValues]["pricePerUnit"])

		storageString = str(storagePrice)
		#print (storageString)

		storageUnitPrice = re.search( r'(\d{1,10}\.\d{1,10})', storageString, re.S)
		#print (storageUnitPrice)

		if storageUnitPrice:
			storageUnitPrice = storageUnitPrice.group(1)
			ws.cell(row=(x), column=11).value = (storageUnitPrice)
			#print ("Storage Unit Price:    " + (storageUnitPrice))

			storageDailyPrice = ((float(storageUnitPrice) * 86400) / (86400 * 30))
			#storageDailyPrice = round(storageDailyPrice,10)
			volumeSizeGB = ((ws.cell(row=x, column=4).value) / 1024)
			storageDailyPrice = (storageDailyPrice * volumeSizeGB)
			ws.cell(row=(x), column=12).value = (storageDailyPrice)
			#print ("Storage Daily Price:   " + str(storageDailyPrice))
		else:
			print ("No storage price match!")

if instanceVar == 'all':
	rowRange()
	awsPricing()
else:
	rowRange()
	assignInstance()

wb.save(args.file)